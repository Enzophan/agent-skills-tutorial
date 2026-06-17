#!/usr/bin/env python3
import pathlib, openpyxl, yaml, sys
from openpyxl.utils import get_column_letter

def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: to_excel.py <questions_folder> <output_folder>")
    src = pathlib.Path(sys.argv[1])
    out_dir = pathlib.Path(sys.argv[2])
    # Ensure the output directory exists (create parents if needed)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "Quiz_updated.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz"
    headers = ["id","topic","difficulty","format","question","options","answer","explanation"]
    ws.append(headers)

    for yaml_file in src.glob("*.yml"):
        rows = yaml.safe_load(yaml_file.read_text()) or []
        for q in rows:
            # Convert options to pipe‑separated string for readability
            opts = q.get("options", [])
            if isinstance(opts, list):
                opts_str = "|".join(opts)
            else:
                opts_str = str(opts)
            ws.append([
                q.get("id"),
                q.get("topic"),
                q.get("difficulty"),
                q.get("format"),
                q.get("question"),
                opts_str,
                q.get("answer"),
                q.get("explanation")
            ])

    # Auto‑size columns (simple heuristic)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    wb.save(out_path)
    print(f"✅ Excel file written to {out_path}")

if __name__ == "__main__":
    main()
