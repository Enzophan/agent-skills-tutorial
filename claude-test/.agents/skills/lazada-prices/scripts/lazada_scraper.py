import sys
import re
import os
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright


def search_lazada(query):
    results = []
    url = f"https://www.lazada.vn/catalog/?q={query.replace(' ', '+')}"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(url, wait_until="networkidle", timeout=30000)
        page.wait_for_timeout(3000)

        content = page.content()
        browser.close()

    soup = BeautifulSoup(content, "lxml")
    items = soup.select("[data-item-id]")

    for item in items:
        title_el = item.select_one(".RfADT a")
        if not title_el:
            title_el = item.select_one(".RfADt a")
        title = title_el.text.strip() if title_el else ""

        price_el = item.select_one(".aBrP0 .ooOxS")
        price_text = price_el.text.strip() if price_el else ""

        discount_el = item.select_one(".WNoq3 .IcOsH")
        discount = discount_el.text.strip() if discount_el else ""

        location_el = item.select_one(".oa6ri")
        location = location_el.get("title", location_el.text.strip()) if location_el else ""

        sales_el = item.select_one("._1cEkb span")
        sales = sales_el.text.strip() if sales_el else ""

        rating_el = item.select_one(".mdmmT span")
        rating = rating_el.text.strip() if rating_el else ""

        link_el = item.select_one(".RfADT a") or item.select_one(".RfADt a")
        link = "https:" + link_el.get("href", "") if link_el else ""

        clean_title = re.sub(r'^\[.*?\]\s*', '', title).strip()

        results.append({
            "title": clean_title,
            "price": price_text,
            "discount": discount,
            "location": location,
            "sales": sales,
            "rating": rating,
            "url": link,
        })

    return results


def export_excel(results, query, filepath):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Lazada Prices"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E8001A", end_color="E8001A", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    headers = ["#", "Product Name", "Price", "Discount", "Location", "Sold", "Rating", "Link"]
    col_widths = [4, 50, 18, 15, 20, 12, 10, 80]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col_idx)].width = width

    for i, r in enumerate(results, 1):
        ws.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=i + 1, column=2, value=r["title"])
        ws.cell(row=i + 1, column=3, value=r["price"])
        ws.cell(row=i + 1, column=4, value=r["discount"])
        ws.cell(row=i + 1, column=5, value=r["location"])
        ws.cell(row=i + 1, column=6, value=r["sales"])
        ws.cell(row=i + 1, column=7, value=r["rating"])
        link_cell = ws.cell(row=i + 1, column=8, value=r["url"])
        link_cell.font = link_font
        link_cell.hyperlink = r["url"]

    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    return os.path.abspath(filepath)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python lazada_scraper.py <product_name> [--excel]")
        print("  --excel    Also export results to an Excel (.xlsx) file")
        sys.exit(1)

    args = sys.argv[1:]
    export_xlsx = "--excel" in args
    if export_xlsx:
        args.remove("--excel")

    query = " ".join(args)
    results = search_lazada(query)

    if not results:
        print(f"No results found for '{query}'")
        sys.exit(0)

    print(f"\n=== Kết quả tìm kiếm cho: {query} ===\n")
    for i, r in enumerate(results, 1):
        print(f"{i}. {r['title']}")
        print(f"   Giá: {r['price']}")
        if r["discount"]:
            print(f"   Giảm: {r['discount']}")
        print(f"   Vị trí: {r['location']}")
        if r["sales"]:
            print(f"   Đã bán: {r['sales']}")
        if r["rating"]:
            print(f"   Đánh giá: {r['rating']}/5")
        print(f"   Link: {r['url']}")
        print()

    if export_xlsx:
        safe_name = re.sub(r'[^\w\s]', '', query).strip().replace(' ', '_')
        xlsx_path = f"lazada_prices_{safe_name}.xlsx"
        full_path = export_excel(results, query, xlsx_path)
        print(f"📊 Exported to Excel: {full_path}")
